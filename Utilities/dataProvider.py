import openpyxl


def get_data(sheetName):
    workbook = openpyxl.load_workbook("D:\PageObjectModelSelenium\Excels\Test data Selenium.xlsx")

    sheet = workbook[sheetName]
    totalrows = sheet.max_row
    totalcols = sheet.max_column
    mainList = []
    print("total rows are:", str(totalrows), "and total Cols are:", str(totalcols))

    for i in range(2, totalrows + 1):
        dataList = []
        for j in range(1, totalcols + 1):
            data = sheet.cell(row=i, column=j).value
            dataList.insert(j, data)
        mainList.insert(i, dataList)
    return mainList
