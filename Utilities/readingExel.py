import openpyxl

workbook = openpyxl.load_workbook("../Excels/Test data Selenium.xlsx")

sheet = workbook["LoginTest"]
totalrows = sheet.max_row
totalcols = sheet.max_colum

print("total rows are:",str(totalrows),"and total Cols are:", str(totalcols))

for rows in range(1,totalrows+1):
    for cols in range(1,totalcols+1):
        print(sheet.cell(row=rows,col=cols).value)


